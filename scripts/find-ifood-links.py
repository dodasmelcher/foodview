#!/usr/bin/env python3
"""find-ifood-links.py — preenche a coluna Link_iFood (e Confianca) de um XLSX
de restaurantes, buscando o site oficial do iFood de cada um por busca na web
e validando nome + endereço.

Engines de busca:
  --engine ddg     (padrão) DuckDuckGo, gratuito, sem chave. Ideal pro primeiro
                   run. Pode receber 'rate-limit' depois de muitas queries —
                   nesse caso ele anota Confianca=0 e segue.
  --engine google  Google Programmable Search Engine (Custom Search JSON API).
                   Precisa criar a chave no console.cloud.google.com e o CX no
                   programmablesearchengine.google.com. Exporte:
                       export GOOGLE_API_KEY=...
                       export GOOGLE_CX=...
                   Gratuito até 100 queries/dia.

Uso:
    pip install -r scripts/requirements-ifood.txt
    python3 scripts/find-ifood-links.py restaurantes.xlsx
    # gera restaurantes-ifood.xlsx ao lado

Re-execução:
    Linhas que já tenham Link_iFood preenchido são puladas — dá pra parar com
    Ctrl-C e retomar depois apontando pra mesma saída.

Confiança (0–100):
    Combina (a) fuzzy match entre o slug do restaurante na URL e o nome do
    lugar, (b) bairro presente no slug, (c) bairro/rua presentes na página do
    iFood (quando --no-verify não está ligado). Sem nenhum sinal de endereço o
    score do nome é capado em 70 — 100 só com confirmação por endereço.
"""

import argparse
import os
import random
import re
import sys
import time
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup
from rapidfuzz import fuzz

# ---------- normalização ----------
_ACCENTS = str.maketrans(
    "áàâãäéèêëíìîïóòôõöúùûüçÁÀÂÃÄÉÈÊËÍÌÎÏÓÒÔÕÖÚÙÛÜÇ",
    "aaaaaeeeeiiiiooooouuuucAAAAAEEEEIIIIOOOOOUUUUC",
)
_SLUG_RE = re.compile(r"[^a-z0-9]+")


def normalize(s: str) -> str:
    return (s or "").translate(_ACCENTS).lower()


def slugify(s: str) -> str:
    return _SLUG_RE.sub("-", normalize(s)).strip("-")


def extract_neighborhood(address: str) -> str:
    """Pega o bairro de um endereço estilo 'Rua X, 123 - Bairro - São Paulo - SP, CEP'."""
    if not address:
        return ""
    parts = [p.strip() for p in re.split(r"\s*-\s*", address) if p.strip()]
    return parts[1] if len(parts) >= 2 else ""


def extract_street(address: str) -> str:
    if not address:
        return ""
    m = re.match(r"^([^,]+)", address)
    return m.group(1).strip() if m else ""


# ---------- engines ----------
def search_ddg(query: str, max_results: int = 5):
    """DuckDuckGo via lib `ddgs` (sucessora de `duckduckgo-search`)."""
    try:
        from ddgs import DDGS
    except ImportError:
        from duckduckgo_search import DDGS  # fallback p/ versão antiga
    with DDGS() as ddgs:
        return [r.get("href") for r in ddgs.text(query, max_results=max_results) if r.get("href")]


def search_google_cse(query: str, max_results: int = 5):
    api_key = os.environ.get("GOOGLE_API_KEY")
    cx = os.environ.get("GOOGLE_CX")
    if not (api_key and cx):
        raise SystemExit(
            "GOOGLE_API_KEY e GOOGLE_CX não configurados — use --engine ddg ou "
            "exporte essas variáveis. Veja o docstring do arquivo."
        )
    r = requests.get(
        "https://www.googleapis.com/customsearch/v1",
        params={"key": api_key, "cx": cx, "q": query, "num": max_results},
        timeout=15,
    )
    r.raise_for_status()
    return [item["link"] for item in r.json().get("items", [])]


# ---------- iFood ----------
IFOOD_RE = re.compile(r"^https?://(?:www\.)?ifood\.com\.br/(?:delivery|restaurante)/", re.I)


def pick_ifood(urls):
    for u in urls:
        if u and IFOOD_RE.match(u):
            return u
    return None


def url_slug(url: str) -> str:
    """Penúltimo segmento da URL de loja do iFood (antes do uuid)."""
    parts = [p for p in urlparse(url).path.split("/") if p]
    if len(parts) >= 3:
        return parts[-2]
    return parts[-1] if parts else ""


# UA realista — sem isso a iFood costuma devolver página vazia.
BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_0) AppleWebKit/605.1.15 "
    "(KHTML, like Gecko) Version/17.0 Safari/605.1.15"
)


def fetch_page_text(url: str, session: requests.Session) -> str:
    try:
        r = session.get(url, timeout=15, headers={"User-Agent": BROWSER_UA, "Accept-Language": "pt-BR,pt;q=0.9"})
        if not r.ok:
            return ""
        return normalize(BeautifulSoup(r.text, "html.parser").get_text(" ", strip=True))
    except Exception:
        return ""


def confidence(name: str, address: str, url: str, page_norm: str) -> int:
    """Score 0–100. Sem confirmação de endereço, fica capado em 70."""
    slug = url_slug(url).replace("-", " ")
    name_norm = slugify(name).replace("-", " ")
    name_score = fuzz.token_set_ratio(name_norm, slug)  # 0..100

    bairro = slugify(extract_neighborhood(address)).replace("-", " ")
    street = slugify(extract_street(address)).replace("-", " ")

    bairro_in_slug = bool(bairro) and bairro in slug
    bairro_in_page = bool(bairro) and bairro in page_norm
    street_in_page = bool(street) and len(street) > 6 and street in page_norm

    score = int(round(name_score * 0.7))  # nome sozinho → até 70
    if bairro_in_slug:
        score += 15
    if bairro_in_page:
        score += 12
    if street_in_page:
        score += 8
    return max(0, min(100, score))


# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="Preenche Link_iFood em um XLSX de restaurantes.")
    ap.add_argument("input", help="XLSX de entrada (colunas: Nome, Endereco, Tipo)")
    ap.add_argument("-o", "--output", default=None, help="XLSX de saída (default: <input>-ifood.xlsx)")
    ap.add_argument("--engine", choices=["ddg", "google"], default="ddg")
    ap.add_argument("--all-types", action="store_true",
                    help="Processa restaurantes E bares (default: só restaurantes).")
    ap.add_argument("--limit", type=int, default=0, help="Limita a N linhas (útil pra teste).")
    ap.add_argument("--delay-min", type=float, default=2.0)
    ap.add_argument("--delay-max", type=float, default=5.0)
    ap.add_argument("--no-verify", action="store_true",
                    help="Não baixa a página do iFood pra confirmar endereço (mais rápido, menos preciso).")
    args = ap.parse_args()

    input_path = Path(args.input)
    out_path = Path(args.output) if args.output else input_path.with_name(input_path.stem + "-ifood.xlsx")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def col(name: str) -> int:
        if name in headers:
            return headers.index(name) + 1
        ws.cell(row=1, column=len(headers) + 1, value=name)
        headers.append(name)
        return len(headers)

    nome_c = col("Nome")
    ender_c = col("Endereco")
    tipo_c = col("Tipo")
    link_c = col("Link_iFood")
    conf_c = col("Confianca")

    search = search_ddg if args.engine == "ddg" else search_google_cse
    session = requests.Session()

    total = ws.max_row - 1
    processed = 0
    found = 0
    print(f"[info] {total} linhas | engine={args.engine} | "
          f"all_types={args.all_types} | verify={'no' if args.no_verify else 'yes'}", file=sys.stderr)

    for row in range(2, ws.max_row + 1):
        if args.limit and processed >= args.limit:
            break
        nome = ws.cell(row=row, column=nome_c).value
        end = ws.cell(row=row, column=ender_c).value or ""
        tipo = (ws.cell(row=row, column=tipo_c).value or "").lower()
        existing = ws.cell(row=row, column=link_c).value

        if existing:
            continue
        if not args.all_types and tipo == "bar":
            continue
        if not nome:
            continue

        bairro = extract_neighborhood(end)
        query = f'site:ifood.com.br "{nome}" {bairro} são paulo'.strip()

        try:
            urls = search(query, max_results=5)
        except Exception as e:
            print(f"[warn] {nome!r}: busca falhou ({e})", file=sys.stderr)
            urls = []

        chosen = pick_ifood(urls)
        conf = 0
        if chosen:
            page_norm = "" if args.no_verify else fetch_page_text(chosen, session)
            conf = confidence(nome, end, chosen, page_norm)
            ws.cell(row=row, column=link_c, value=chosen)
            ws.cell(row=row, column=conf_c, value=conf)
            found += 1
        processed += 1
        print(f"[{processed}/{total}] {nome}  →  {chosen or '—'}  ({conf})", file=sys.stderr)

        # Salva incremental a cada 10 pra sobreviver a Ctrl-C / rate limit.
        if processed % 10 == 0:
            wb.save(out_path)

        time.sleep(random.uniform(args.delay_min, args.delay_max))

    wb.save(out_path)
    print(f"[done] {found}/{processed} com link  →  {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
