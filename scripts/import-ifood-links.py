#!/usr/bin/env python3
"""import-ifood-links.py — lê um XLSX com colunas (id, Nome, Link_iFood) e
preenche places.delivery_url no Supabase pra cada linha com link.

Requer:
    export SUPABASE_SERVICE_KEY=...
    python3 scripts/import-ifood-links.py restaurantes-ifood.xlsx

Efeito por linha (só onde Link_iFood não tá vazio):
  • delivery_url    ← link
  • has_delivery    ← true
  • delivery_apps   ← preserva os apps que já estavam lá e adiciona "iFood"
                      se ainda não estiver (case-insensitive)
"""

import os
import sys
from pathlib import Path

import openpyxl
import requests

URL = "https://jspxkdhqhjjvtepomkir.supabase.co"
KEY = os.environ.get("SUPABASE_SERVICE_KEY")
if not KEY:
    sys.exit("ERRO: defina SUPABASE_SERVICE_KEY no ambiente.")

if len(sys.argv) < 2:
    sys.exit("uso: import-ifood-links.py <xlsx>")

wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active
h = [c.value for c in ws[1]]
try:
    i_id = h.index("id") + 1
    i_link = h.index("Link_iFood") + 1
    i_nome = h.index("Nome") + 1
except ValueError as e:
    sys.exit(f"ERRO: planilha precisa das colunas id, Nome, Link_iFood ({e})")

HEADERS = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

# Pega delivery_apps atuais de todos os ids em lote pra evitar 1 GET por linha.
ids_with_link = [
    (ws.cell(row=r, column=i_id).value,
     ws.cell(row=r, column=i_link).value,
     ws.cell(row=r, column=i_nome).value)
    for r in range(2, ws.max_row + 1)
    if ws.cell(row=r, column=i_link).value and ws.cell(row=r, column=i_id).value
]
if not ids_with_link:
    sys.exit("ERRO: planilha não tem nenhuma linha com Link_iFood preenchido.")

id_list = ",".join(str(pid) for pid, _, _ in ids_with_link)
res = requests.get(
    f"{URL}/rest/v1/places?id=in.({id_list})&select=id,delivery_apps",
    headers={"apikey": KEY, "Authorization": f"Bearer {KEY}"},
    timeout=30,
)
res.raise_for_status()
current = {row["id"]: (row.get("delivery_apps") or "") for row in res.json()}

updated, errors = 0, 0
for pid, link, nome in ids_with_link:
    apps_raw = current.get(pid, "")
    apps_list = [a.strip() for a in apps_raw.split(",") if a.strip()]
    if not any(a.lower() == "ifood" for a in apps_list):
        apps_list.append("iFood")
    body = {
        "delivery_url": link,
        "has_delivery": True,
        "delivery_apps": ",".join(apps_list),
    }
    r = requests.patch(
        f"{URL}/rest/v1/places?id=eq.{pid}", headers=HEADERS, json=body, timeout=20
    )
    if r.ok:
        updated += 1
        print(f"  [{updated}/{len(ids_with_link)}] {nome}")
    else:
        errors += 1
        print(f"  [ERR] {nome} (id={pid}): {r.status_code} {r.text[:160]}", file=sys.stderr)

print(f"\nfeito: {updated} atualizadas, {errors} erros.")
